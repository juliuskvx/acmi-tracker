#!/usr/bin/env python3
"""
export_excel.py — ACMI Tracker daily Excel logger (v3)
- Appends one row per aircraft per day to data sheets (deduped by report_date)
- Rebuilds Dashboard sheet from ALL accumulated history on every run
- AutoFilter + date column on every data sheet for Excel date-range filtering
"""

import json, os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "../data")
SNAPSHOT = os.path.join(DATA_DIR, "acmi_data.json")
HISTORY  = os.path.join(DATA_DIR, "acmi_history.json")
EXCEL    = os.path.join(DATA_DIR, "acmi_log.xlsx")

# ── Palette ───────────────────────────────────────────────────────────────────
C_BG       = "060910"
C_HEADER   = "0D1320"
C_ROW_ALT  = "0A1220"
C_CYAN     = "00D4FF"
C_GREEN    = "00CC66"
C_AMBER    = "FFAA00"
C_MUTED    = "64748B"
C_WHITE    = "E2E8F0"
C_PURPLE   = "A78BFA"
C_DARK_ROW = "111827"
C_SEP      = "0A1628"

def _f(sz=9, bold=False, color="E2E8F0", italic=False):
    return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)

def _fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

def _border(color="1E2D42"):
    return Border(bottom=Side(style="thin", color=color))

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = _fill(C_HEADER)
        c.font      = _f(bold=True, color=C_CYAN)
        c.alignment = CENTER
        c.border    = _border("2D4A6B")

def row_bg(i):
    return C_ROW_ALT if i % 2 else C_BG

# ── Dedup helper ──────────────────────────────────────────────────────────────

def existing_report_dates(ws, date_col=2):
    """Return set of report_date strings already in the sheet (column date_col, skip row 1)."""
    dates = set()
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col, values_only=True):
        v = row[0]
        if v and not str(v).startswith("──"):
            dates.add(str(v)[:10])
    return dates

# ── Sheet 1: Snapshot Log ─────────────────────────────────────────────────────

SNAP_HEADERS = [
    "Date", "Report Date", "Registration", "Type", "Owner ICAO", "Owner Name",
    "Group", "Status", "Callsign", "Operator ICAO", "Operator Name",
    "Route", "Last Seen", "Altitude", "Speed", "Heading",
]
SNAP_WIDTHS = [11, 11, 12, 7, 11, 20, 20, 13, 13, 13, 22, 13, 18, 9, 7, 7]

def append_snapshot(ws, snap):
    is_new = ws.max_row == 1 and ws.cell(1, 1).value is None
    ws.sheet_view.showGridLines = False

    if is_new:
        ws.append(SNAP_HEADERS)
        style_header_row(ws, 1, len(SNAP_HEADERS))
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(SNAP_HEADERS))}1"
        set_col_widths(ws, SNAP_WIDTHS)
        ws.row_dimensions[1].height = 18

    rep_date = snap.get("report_date", "")
    if rep_date in existing_report_dates(ws):
        print(f"  Snapshot: {rep_date} already in sheet — skipping")
        return

    run_date = snap.get("last_updated", "")[:10]
    fleet    = snap.get("fleet", [])

    for i, ac in enumerate(fleet):
        status = ac.get("status", "")
        r  = ws.max_row + 1
        bg = row_bg(i)
        ws.row_dimensions[r].height = 14

        row_data = [
            run_date, rep_date,
            ac.get("registration"), ac.get("type"),
            ac.get("owner_icao"),   ac.get("owner_name"), ac.get("group"),
            ac.get("status"),       ac.get("callsign"),
            ac.get("current_operator_icao"), ac.get("current_operator_name"),
            ac.get("last_route"),
            ac.get("last_seen", "")[:19].replace("T", " ") if ac.get("last_seen") else None,
            ac.get("altitude"), ac.get("speed"), ac.get("heading"),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(r, col, val)
            c.fill      = _fill(bg)
            c.font      = _f(color=C_WHITE)
            c.alignment = LEFT
            c.border    = _border()

        if status == "acmi_active":
            ws.cell(r, 3).font  = _f(bold=True, color=C_CYAN)
            ws.cell(r, 8).font  = _f(bold=True, color=C_CYAN)
            ws.cell(r, 11).font = _f(color=C_GREEN)
        elif status == "own_ops":
            ws.cell(r, 8).font = _f(color=C_AMBER)
        else:
            ws.cell(r, 3).font = _f(color=C_MUTED)
            ws.cell(r, 8).font = _f(color=C_MUTED)

# ── Sheet 2: Daily Report Log ─────────────────────────────────────────────────

HIST_HEADERS = [
    "Date", "Report Date", "Registration", "Type", "Owner ICAO", "Owner Name", "Group",
    "Total Flights", "Total BH", "ACMI Flights", "ACMI BH",
    "Clients (name · BH)", "Routes",
]
HIST_WIDTHS = [11, 11, 12, 7, 11, 20, 20, 12, 10, 12, 10, 44, 36]

def append_history(ws, hist):
    is_new = ws.max_row == 1 and ws.cell(1, 1).value is None
    ws.sheet_view.showGridLines = False

    if is_new:
        ws.append(HIST_HEADERS)
        style_header_row(ws, 1, len(HIST_HEADERS))
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HIST_HEADERS))}1"
        set_col_widths(ws, HIST_WIDTHS)
        ws.row_dimensions[1].height = 18

    rep_date = hist.get("report_date", "")
    if rep_date in existing_report_dates(ws):
        print(f"  Daily Report: {rep_date} already in sheet — skipping")
        return

    run_date = hist.get("last_updated", "")[:10]
    fleet    = hist.get("fleet", [])

    active = sorted(
        [a for a in fleet if a.get("total_flights", 0) > 0],
        key=lambda x: -x.get("total_bh", 0)
    )

    for i, ac in enumerate(active):
        r  = ws.max_row + 1
        bg = row_bg(i)
        ws.row_dimensions[r].height = 14

        tbh = round(ac.get("total_bh", 0) or 0, 1)
        abh = round(ac.get("acmi_bh",  0) or 0, 1)

        clients_str = "  |  ".join(
            f"{c.get('name') or c.get('icao')} · {c.get('bh', 0)}h"
            for c in (ac.get("clients") or [])
        )
        routes_str = "  ".join(ac.get("routes") or [])

        row_data = [
            run_date, rep_date,
            ac.get("registration"), ac.get("type"),
            ac.get("owner_icao"),   ac.get("owner_name"), ac.get("group"),
            ac.get("total_flights"), tbh,
            ac.get("acmi_flights") or 0, abh,
            clients_str or None, routes_str or None,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(r, col, val)
            c.fill      = _fill(bg)
            c.font      = _f(color=C_WHITE)
            c.alignment = LEFT
            c.border    = _border()

        for col in [8, 9, 10, 11]:
            ws.cell(r, col).alignment = RIGHT

        if ac.get("acmi_flights", 0):
            ws.cell(r, 3).font  = _f(bold=True, color=C_CYAN)
            ws.cell(r, 11).font = _f(bold=True, color=C_GREEN)
        else:
            ws.cell(r, 3).font = _f(color=C_MUTED)

        if tbh > 0:
            ws.cell(r, 9).font = _f(color=C_GREEN)

# ── Sheet 3: Operator Summary ─────────────────────────────────────────────────

OPS_HEADERS = [
    "Date", "Report Date", "Operator Group",
    "Total Flights", "Total BH", "ACMI Flights", "ACMI BH", "ACMI %", "Active Aircraft",
]
OPS_WIDTHS = [11, 11, 22, 13, 10, 13, 10, 9, 15]

def append_op_summaries(ws, hist):
    is_new = ws.max_row == 1 and ws.cell(1, 1).value is None
    ws.sheet_view.showGridLines = False

    if is_new:
        ws.append(OPS_HEADERS)
        style_header_row(ws, 1, len(OPS_HEADERS))
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(OPS_HEADERS))}1"
        set_col_widths(ws, OPS_WIDTHS)
        ws.row_dimensions[1].height = 18

    rep_date = hist.get("report_date", "")
    if rep_date in existing_report_dates(ws):
        print(f"  Operator Summary: {rep_date} already in sheet — skipping")
        return

    run_date = hist.get("last_updated", "")[:10]
    ops      = hist.get("operator_summaries", {})

    for i, (group, data) in enumerate(
        sorted(ops.items(), key=lambda x: -x[1].get("acmi_bh", 0))
    ):
        r  = ws.max_row + 1
        bg = row_bg(i)
        ws.row_dimensions[r].height = 14

        tbh  = round(data.get("total_bh", 0) or 0, 1)
        abh  = round(data.get("acmi_bh",  0) or 0, 1)
        rate = round(abh / tbh * 100, 1) if tbh else 0.0

        row_data = [
            run_date, rep_date, group,
            data.get("total_flights"), tbh,
            data.get("acmi_flights"),  abh,
            rate / 100,   # stored as fraction; formatted as %
            data.get("active_aircraft"),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(r, col, val)
            c.fill      = _fill(bg)
            c.font      = _f(color=C_WHITE)
            c.alignment = LEFT
            c.border    = _border()

        ws.cell(r, 8).number_format = "0.0%"
        for col in [4, 5, 6, 7, 8, 9]:
            ws.cell(r, col).alignment = RIGHT

        if abh > 0:
            ws.cell(r, 3).font = _f(bold=True, color=C_CYAN)
            ws.cell(r, 7).font = _f(bold=True, color=C_GREEN)
        if tbh > 0:
            ws.cell(r, 5).font = _f(color=C_WHITE)

        rate_color = C_GREEN if rate >= 80 else (C_AMBER if rate >= 40 else C_MUTED)
        ws.cell(r, 8).font = _f(bold=(rate >= 80), color=rate_color)

# ── Sheet 4: Client Map ───────────────────────────────────────────────────────

CLIENT_HEADERS = [
    "Date", "Report Date", "Client ICAO", "Client Name",
    "ACMI Provider", "Flights", "BH", "Aircraft Count", "Aircraft Regs",
]
CLIENT_WIDTHS = [11, 11, 13, 24, 20, 9, 9, 13, 44]

def append_client_map(ws, hist):
    is_new = ws.max_row == 1 and ws.cell(1, 1).value is None
    ws.sheet_view.showGridLines = False

    if is_new:
        ws.append(CLIENT_HEADERS)
        style_header_row(ws, 1, len(CLIENT_HEADERS))
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(CLIENT_HEADERS))}1"
        set_col_widths(ws, CLIENT_WIDTHS)
        ws.row_dimensions[1].height = 18

    rep_date = hist.get("report_date", "")
    if rep_date in existing_report_dates(ws):
        print(f"  Client Map: {rep_date} already in sheet — skipping")
        return

    run_date = hist.get("last_updated", "")[:10]
    cm       = hist.get("client_map", {})
    if not cm:
        return

    def client_total_bh(item):
        _, val = item
        providers = val["providers"] if isinstance(val, dict) and "providers" in val else val
        return sum(p.get("bh", 0) for p in providers.values())

    row_idx = 0
    for client_icao, val in sorted(cm.items(), key=client_total_bh, reverse=True):
        if isinstance(val, dict) and "providers" in val:
            client_name = val.get("name", client_icao)
            providers   = val["providers"]
        else:
            client_name = client_icao
            providers   = val

        for provider, data in sorted(providers.items(), key=lambda x: -x[1].get("bh", 0)):
            r  = ws.max_row + 1
            bg = row_bg(row_idx)
            bh = round(data.get("bh", 0) or 0, 1)
            ws.row_dimensions[r].height = 14

            row_data = [
                run_date, rep_date,
                client_icao, client_name,
                provider,
                data.get("flights"), bh,
                data.get("aircraft_count"),
                ", ".join(data.get("aircraft") or []),
            ]
            for col, val2 in enumerate(row_data, 1):
                c = ws.cell(r, col, val2)
                c.fill      = _fill(bg)
                c.font      = _f(color=C_WHITE)
                c.alignment = LEFT
                c.border    = _border()

            for col in [6, 7, 8]:
                ws.cell(r, col).alignment = RIGHT

            ws.cell(r, 3).font = _f(bold=True, color=C_PURPLE)
            ws.cell(r, 4).font = _f(bold=True, color=C_WHITE)
            ws.cell(r, 5).font = _f(color=C_CYAN)
            for col in [6, 7, 8]:
                ws.cell(r, col).alignment = RIGHT
            if bh > 0:
                ws.cell(r, 7).font = _f(bold=True, color=C_GREEN)

            row_idx += 1

# ── Dashboard: rebuild from ALL accumulated data ──────────────────────────────

def build_dashboard(ws, ops_ws, dr_ws, cm_ws):
    """Read accumulated data from the other sheets and render a summary Dashboard."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00D4FF"

    def is_date(v):
        """Return True only for real YYYY-MM-DD date strings, skip headers/separators."""
        if not v:
            return False
        s = str(v)[:10]
        return len(s) == 10 and s[4] == "-" and s[7] == "-" and s[:4].isdigit()

    # ── Read Operator Summary sheet ──
    ops_data = []
    for row in ops_ws.iter_rows(min_row=1, values_only=True):
        if not is_date(row[1]) or not row[2]:
            continue
        try:
            ops_data.append({
                "report_date":  str(row[1])[:10],
                "operator":     row[2],
                "tot_flights":  int(row[3]  or 0),
                "tot_bh":       float(row[4] or 0),
                "acmi_flights": int(row[5]  or 0),
                "acmi_bh":      float(row[6] or 0),
                "active_ac":    int(row[7]  or 0) if len(row) > 7 else 0,
            })
        except (ValueError, TypeError):
            continue

    # ── Read Client Map sheet ──
    cm_data = []
    for row in cm_ws.iter_rows(min_row=1, values_only=True):
        if not is_date(row[1]) or not row[2]:
            continue
        try:
            cm_data.append({
                "report_date": str(row[1])[:10],
                "client_icao": row[2],
                "client_name": row[3] or row[2],
                "provider":    row[4],
                "flights":     int(row[5]  or 0),
                "bh":          float(row[6] or 0),
                "aircraft":    int(row[7]  or 0) if len(row) > 7 else 0,
            })
        except (ValueError, TypeError):
            continue

    # ── Aggregate ALL-TIME totals ──
    all_dates = sorted(set(r["report_date"] for r in ops_data))
    latest    = all_dates[-1] if all_dates else "N/A"
    earliest  = all_dates[0]  if all_dates else "N/A"
    n_days    = len(all_dates)

    def ops_sum(data, field):
        return round(sum(r[field] for r in data), 1)

    all_tot_bh   = ops_sum(ops_data, "tot_bh")
    all_acmi_bh  = ops_sum(ops_data, "acmi_bh")
    all_flights  = ops_sum(ops_data, "tot_flights")
    all_acmi_fl  = ops_sum(ops_data, "acmi_flights")
    all_clients  = len(set(r["client_icao"] for r in cm_data))
    all_ops      = len(set(r["operator"]    for r in ops_data))
    acmi_rate    = round(all_acmi_bh / all_tot_bh * 100, 1) if all_tot_bh else 0

    # Latest-day stats
    latest_ops  = [r for r in ops_data  if r["report_date"] == latest]
    latest_cm   = [r for r in cm_data   if r["report_date"] == latest]
    lat_bh      = round(sum(r["tot_bh"]   for r in latest_ops), 1)
    lat_acmi_bh = round(sum(r["acmi_bh"]  for r in latest_ops), 1)
    lat_flights = sum(r["tot_flights"]     for r in latest_ops)
    lat_rate    = round(lat_acmi_bh / lat_bh * 100, 1) if lat_bh else 0

    # Per-operator all-time totals
    from collections import defaultdict
    op_totals = defaultdict(lambda: {"tot_bh": 0, "acmi_bh": 0, "tot_flights": 0, "acmi_flights": 0, "days": 0})
    for r in ops_data:
        op = r["operator"]
        op_totals[op]["tot_bh"]      += r["tot_bh"]
        op_totals[op]["acmi_bh"]     += r["acmi_bh"]
        op_totals[op]["tot_flights"] += r["tot_flights"]
        op_totals[op]["acmi_flights"]+= r["acmi_flights"]
        op_totals[op]["days"]        += 1

    # Per-client all-time totals
    cl_totals = defaultdict(lambda: {"name": "", "bh": 0, "flights": 0, "providers": set()})
    for r in cm_data:
        cl = r["client_icao"]
        cl_totals[cl]["name"]     = r["client_name"]
        cl_totals[cl]["bh"]      += r["bh"]
        cl_totals[cl]["flights"] += r["flights"]
        cl_totals[cl]["providers"].add(r["provider"])

    # ── Render ──
    def put(row, col, val, color=C_WHITE, sz=9, bold=False, align=LEFT, fill=C_BG, italic=False, num_fmt=None):
        c = ws.cell(row, col, val)
        c.font      = _f(sz=sz, bold=bold, color=color, italic=italic)
        c.fill      = _fill(fill)
        c.alignment = align
        if num_fmt:
            c.number_format = num_fmt
        return c

    def section(row, title):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        c = ws.cell(row, 2, title)
        c.font      = _f(bold=True, color=C_AMBER, sz=9)
        c.fill      = _fill(C_DARK_ROW)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 18

    def kpi_block(label_row, val_row, items):
        for col, label, val, color in items:
            ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=col+1)
            lc = ws.cell(label_row, col, label)
            lc.font      = _f(sz=8, bold=True, color=C_MUTED)
            lc.fill      = _fill(C_HEADER)
            lc.alignment = CENTER

            ws.merge_cells(start_row=val_row, start_column=col, end_row=val_row, end_column=col+1)
            vc = ws.cell(val_row, col, val)
            vc.font      = _f(sz=20, bold=True, color=color)
            vc.fill      = _fill(C_HEADER)
            vc.alignment = CENTER
            vc.border    = _border("2D4A6B")

    def tbl_header(row, cols_labels):
        for col, label in cols_labels:
            c = ws.cell(row, col, label)
            c.font      = _f(sz=8, bold=True, color=C_CYAN)
            c.fill      = _fill(C_HEADER)
            c.alignment = CENTER
            c.border    = _border("2D4A6B")

    # Silence all cells with BG first
    for r in range(1, 60):
        ws.row_dimensions[r].height = 15
        for col in range(1, 15):
            c = ws.cell(r, col)
            c.fill = _fill(C_BG)

    # Row 1 spacer
    ws.row_dimensions[1].height = 6

    # Title
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:M2")
    t = ws.cell(2, 2, "✈  ACMI INTEL — DASHBOARD")
    t.font      = _f(sz=18, bold=True, color=C_CYAN)
    t.fill      = _fill(C_HEADER)
    t.alignment = CENTER

    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:M3")
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    sub = ws.cell(3, 2, f"Data: {earliest} → {latest}  ({n_days} days)   |   Refreshed: {now_str}")
    sub.font      = _f(sz=9, color=C_MUTED, italic=True)
    sub.fill      = _fill(C_HEADER)
    sub.alignment = CENTER

    ws.row_dimensions[4].height = 8

    # ── ALL-TIME KPIs ──
    section(5, "ALL-TIME TOTALS")
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 38
    ws.row_dimensions[8].height = 8
    kpi_block(6, 7, [
        (2,  "TOTAL BH",      f"{all_tot_bh:,.0f}h",  C_WHITE),
        (4,  "ACMI BH",       f"{all_acmi_bh:,.0f}h", C_GREEN),
        (6,  "ACMI RATE",     f"{acmi_rate}%",         C_GREEN),
        (8,  "TOTAL FLIGHTS", f"{int(all_flights):,}", C_WHITE),
        (10, "ACMI FLIGHTS",  f"{int(all_acmi_fl):,}", C_CYAN),
        (12, "CLIENTS SERVED",str(all_clients),        C_PURPLE),
    ])

    # ── LATEST DAY KPIs ──
    section(9, f"LATEST DAY — {latest}")
    ws.row_dimensions[10].height = 14
    ws.row_dimensions[11].height = 38
    ws.row_dimensions[12].height = 8
    kpi_block(10, 11, [
        (2,  "TOTAL BH",     f"{lat_bh}h",        C_WHITE),
        (4,  "ACMI BH",      f"{lat_acmi_bh}h",   C_GREEN),
        (6,  "ACMI RATE",    f"{lat_rate}%",       C_GREEN if lat_rate >= 50 else C_AMBER),
        (8,  "FLIGHTS",      str(lat_flights),     C_WHITE),
        (10, "OPERATORS",    str(len(latest_ops)), C_CYAN),
        (12, "CLIENTS",      str(len(set(r["client_icao"] for r in latest_cm))), C_PURPLE),
    ])

    # ── Operator all-time table ──
    section(13, "OPERATORS — ALL-TIME BLOCK HOURS")
    ws.row_dimensions[14].height = 16
    tbl_header(14, [(2,"OPERATOR"),(5,"TOTAL BH"),(6,"ACMI BH"),(7,"ACMI %"),(8,"FLIGHTS"),(9,"ACMI FLT"),(10,"DAYS")])

    sorted_ops_at = sorted(op_totals.items(), key=lambda x: -x[1]["acmi_bh"])
    row = 15
    for i, (op, d) in enumerate(sorted_ops_at):
        bg   = row_bg(i)
        tbh  = round(d["tot_bh"],  1)
        abh  = round(d["acmi_bh"], 1)
        rate = round(abh / tbh * 100, 1) if tbh else 0
        ws.row_dimensions[row].height = 15
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

        for col, val, color, align in [
            (2,  op,                     C_CYAN if abh > 0 else C_WHITE, LEFT),
            (5,  tbh,                    C_WHITE,                         RIGHT),
            (6,  abh,                    C_GREEN if abh > 0 else C_MUTED, RIGHT),
            (7,  rate / 100,             C_GREEN if rate >= 80 else (C_AMBER if rate >= 40 else C_MUTED), RIGHT),
            (8,  d["tot_flights"],       C_WHITE,  RIGHT),
            (9,  d["acmi_flights"],      C_CYAN,   RIGHT),
            (10, d["days"],              C_MUTED,  RIGHT),
        ]:
            c = ws.cell(row, col, val)
            c.font      = _f(color=color, bold=(col == 2 and abh > 0))
            c.fill      = _fill(bg)
            c.alignment = align
            c.border    = _border()
        ws.cell(row, 7).number_format = "0.0%"
        row += 1

    row += 1

    # ── Top clients all-time table ──
    section(row, "TOP CLIENTS — ALL-TIME ACMI BLOCK HOURS")
    row += 1
    ws.row_dimensions[row].height = 16
    tbl_header(row, [(2,"CLIENT"),(5,"ACMI PROVIDER(S)"),(8,"FLIGHTS"),(9,"BH"),(10,"AIRCRAFT")])
    row += 1

    sorted_clients = sorted(cl_totals.items(), key=lambda x: -x[1]["bh"])
    for i, (icao, d) in enumerate(sorted_clients[:15]):
        bg = row_bg(i)
        ws.row_dimensions[row].height = 15
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)

        for col, val, color, align in [
            (2,  f"{d['name']} ({icao})",  C_PURPLE, LEFT),
            (5,  ", ".join(d["providers"]), C_WHITE,  LEFT),
            (8,  d["flights"],              C_WHITE,  RIGHT),
            (9,  round(d["bh"], 1),         C_GREEN,  RIGHT),
            (10, "",                         C_MUTED,  RIGHT),
        ]:
            c = ws.cell(row, col, val)
            c.font      = _f(color=color, bold=(col == 2))
            c.fill      = _fill(bg)
            c.alignment = align
            c.border    = _border()
        row += 1

    # Column widths
    col_widths = {1:2, 2:14, 3:6, 4:6, 5:12, 6:10, 7:9, 8:10, 9:10, 10:10, 11:10, 12:10, 13:2}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    snap = hist = None

    if os.path.exists(SNAPSHOT):
        with open(SNAPSHOT) as f:
            snap = json.load(f)
        print(f"Loaded snapshot: {snap.get('report_date')} — {len(snap.get('fleet', []))} aircraft")
    else:
        print("WARNING: acmi_data.json not found")

    if os.path.exists(HISTORY):
        with open(HISTORY) as f:
            raw_hist = json.load(f)
        # Support both multi-day {"days": [...]} and legacy single-day format
        if isinstance(raw_hist.get("days"), list) and raw_hist["days"]:
            hist = raw_hist["days"][0]  # newest day first
        else:
            hist = raw_hist
        print(f"Loaded history:  {hist.get('report_date')} — {len(hist.get('fleet', []))} records")
    else:
        print("WARNING: acmi_history.json not found")

    if not snap and not hist:
        print("ERROR: No data files found.")
        raise SystemExit(1)

    # Load existing workbook or create new
    if os.path.exists(EXCEL):
        wb = load_workbook(EXCEL)
        print(f"Loaded existing {EXCEL}")
    else:
        wb = Workbook()
        wb.remove(wb.active)
        print(f"Creating new {EXCEL}")

    def get_or_create(name):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)

    # Append new daily data (deduped)
    if snap:
        append_snapshot(get_or_create("Snapshot Log"), snap)
        print("✓ Snapshot sheet")

    if hist:
        append_history(get_or_create("Daily Report Log"), hist)
        append_op_summaries(get_or_create("Operator Summary"), hist)
        append_client_map(get_or_create("Client Map Log"), hist)
        print("✓ History / Operator / Client sheets")

    # Ensure autofilter on all data sheets (migration for older files)
    sheet_filter_ranges = {
        "Snapshot Log":     f"A1:{get_column_letter(len(SNAP_HEADERS))}1",
        "Daily Report Log": f"A1:{get_column_letter(len(HIST_HEADERS))}1",
        "Operator Summary": f"A1:{get_column_letter(len(OPS_HEADERS))}1",
        "Client Map Log":   f"A1:{get_column_letter(len(CLIENT_HEADERS))}1",
    }
    for sname, ref in sheet_filter_ranges.items():
        if sname in wb.sheetnames and not wb[sname].auto_filter.ref:
            wb[sname].auto_filter.ref = ref
            ws_fix = wb[sname]
            if ws_fix.cell(1, 1).value is None and ws_fix.cell(2, 1).value == "Date":
                ws_fix.delete_rows(1)
            print(f"  ✓ AutoFilter added to {sname}")

    # Rebuild Dashboard from all accumulated data
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)
    build_dashboard(dash, wb["Operator Summary"], wb["Daily Report Log"], wb["Client Map Log"])
    print("✓ Dashboard rebuilt from all history")

    wb.save(EXCEL)
    print(f"\n✓ Saved → {EXCEL}")

if __name__ == "__main__":
    main()
